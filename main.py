import os
from openpyxl import load_workbook as lwb
import openpyxl


def define_exel():
    list_dir = os.listdir()
    for fill_name in list_dir:
        if ".xlsx" in fill_name:
            if "~" not in fill_name:
                name_file_exel = fill_name

    return name_file_exel


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print('Запуск')
    name_xl = define_exel()
    print("Имя файла - ", name_xl)
    wb = lwb(name_xl)
    ws = wb.active

    for i in range(1, ws.max_row + 1):
        x = ws.cell(i, 1).value
        if x is None:
            print("None в строке - ", i)
            for j in range(1, 15):
                ws.cell(i, j).value = ws.cell((i - 1), j).value

    wb.save('as.xlsx')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
